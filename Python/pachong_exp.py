from urllib import request
from urllib import parse
from bs4 import BeautifulSoup
from openpyxl import Workbook

words,classes,sub_group = [],[],[]
url = 'https://diablo4.cc/cn/Affix'

headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:6.0) Gecko/20100101 Firefox/6.0'}

req = request.Request(url=url,headers=headers)
res = request.urlopen(req)
html = res.read().decode("utf-8")
tree = BeautifulSoup(html, 'lxml')

p_tag = tree.select('#Basic词缀 > div > div.collapse.show > div > div > div > div > div.card-header ')
# print(p_tag)
for i in p_tag:
    i = str(i)
    tag = i[i.find('</span>') + 7:i.rfind('<span')].replace('<span class="c_label">幸运一击: </span>','幸运一击:')
    words.append(tag)
# print(words)
print(len(words))

p_tag = tree.select('#Class词缀 > div > div.collapse.show > div > div > div > div > div.card-header ')
# print(p_tag)
for i in p_tag:
    i = str(i)
    tag = i[i.find('</span>') + 7:i.rfind('<span')].replace('<span class="c_label">幸运一击: </span>','幸运一击:')
    words.append(tag)


p_armor = tree.select('div.col > div div:nth-child(2) > div')
for i in p_armor:
    sub_group = []
    # print(i)
    for j in i:
        j = str(j)
        if j.find(',')>=0:
            1
        else:
            sub_group.append(j[j.find('>')+1:j.find('</a>')])
    classes.append(','.join(sub_group))

print(words)
print(classes)
# print(len(words),len(classes))




# 创建一个Workbook对象
wb = Workbook()
# 获取默认的工作表
ws = wb.active
# 指定要写入数据的起始行和列索引
row_num1,row_num2 = 2,2  # 从第2行开始写入数据

# 将列表中的数据写入指定的单元格
for x in words:
    # 计算目标单元格的位置
    cell = ws.cell(row=row_num1, column=1)
    row_num1 += 1
    # 写入数据到单元格
    cell.value = x

# 将列表中的数据写入指定的单元格
for y in classes:
    # 计算目标单元格的位置
    cell = ws.cell(row=row_num2, column=2)
    row_num2 += 1
    # 写入数据到单元格
    cell.value = y

# 保存Excel文件
wb.save('output.xlsx')


